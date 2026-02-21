import io, os, uuid, json, base64
from typing import Optional, Tuple

import numpy as np
from PIL import Image, ImageDraw, ImageFont

import torch
import torch.nn as nn
from torchvision import models, transforms

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse

from fastapi import Body
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import zipfile
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl import Workbook


# ===================== CONFIG =====================

CLASSES = [
    "Algas",
    "Coral",
    "Otros organismos",
    "Sustrato inerte",
    "Tape",
    "nan"
]
NUM_CLASSES = len(CLASSES)

try:
    torch.set_num_threads(4)
except Exception:
    pass

MODEL_CACHE = {}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
MODELS_DIR = os.path.join(BASE_DIR, "models_store")

os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(MODELS_DIR, exist_ok=True)

app = FastAPI()


# ===================== LABELS A..Z, AA.. =====================

def idx_to_label_excel(idx: int) -> str:
    label = ""
    while idx > 0:
        idx -= 1
        label = chr(ord("A") + (idx % 26)) + label
        idx //= 26
    return label


# ===================== POINTS =====================

def generate_random_points(w: int, h: int, n: int, margin: int = 10):
    xs = np.random.randint(margin, max(margin + 1, w - margin), size=n)
    ys = np.random.randint(margin, max(margin + 1, h - margin), size=n)

    points = []
    for i, (x, y) in enumerate(zip(xs, ys), start=1):
        points.append({
            "idx": i,
            "label": idx_to_label_excel(i),
            "x": int(x),
            "y": int(y),
            "x_norm": float(x / w),
            "y_norm": float(y / h),
        })
    return points


# ===================== DRAW =====================

def _load_font(font_size: int):
    for fp in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/Arial.ttf",
        "C:\\Windows\\Fonts\\arial.ttf"
    ]:
        try:
            return ImageFont.truetype(fp, font_size)
        except Exception:
            pass
    return ImageFont.load_default()


def draw_points(img: Image.Image, points):
    out = img.convert("RGB").copy()
    draw = ImageDraw.Draw(out)

    W, H = out.size
    cross_size = max(10, int(min(W, H) * 0.020))
    half = cross_size // 2
    font_size = max(14, int(min(W, H) * 0.020))
    font = _load_font(font_size)

    color = (255, 0, 0)
    width = 2

    for p in points:
        x, y = p["x"], p["y"]
        label = p.get("label", "")

        draw.line((x - half, y, x + half, y), fill=color, width=width)
        draw.line((x, y - half, x, y + half), fill=color, width=width)

        tx, ty = x + half + 4, y - half - 2
        draw.text((tx, ty), label, fill=color, font=font)

    return out


# ===================== MODEL =====================

def build_model(arch: str, num_classes: int):
    arch = arch.lower()

    if arch == "vgg16":
        m = models.vgg16(weights=None)
        m.classifier[6] = nn.Linear(m.classifier[6].in_features, num_classes)
        return m

    if arch == "resnet18":
        m = models.resnet18(weights=None)
        m.fc = nn.Linear(m.fc.in_features, num_classes)
        return m

    if arch == "alexnet":
        m = models.alexnet(weights=None)
        m.classifier[6] = nn.Linear(m.classifier[6].in_features, num_classes)
        return m

    if arch == "mobilenet_v2":
        m = models.mobilenet_v2(weights=None)
        m.classifier[1] = nn.Linear(m.classifier[1].in_features, num_classes)
        return m

    raise ValueError(f"Arquitectura no soportada: {arch}")


def load_model_by_id(model_id: str):
    cached = MODEL_CACHE.get(model_id)
    if cached is not None:
        return cached

    model_dir = os.path.join(MODELS_DIR, model_id)
    meta_path = os.path.join(model_dir, "meta.json")
    pt_path = os.path.join(model_dir, "best_model.pt")

    if not os.path.isfile(meta_path) or not os.path.isfile(pt_path):
        raise FileNotFoundError(f"Modelo '{model_id}' no encontrado")

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    arch = meta.get("model_name", model_id)
    model = build_model(arch, NUM_CLASSES)

    state = torch.load(pt_path, map_location="cpu")
    model.load_state_dict(state)
    model.eval()

    tfm = transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225]),
    ])

    MODEL_CACHE[model_id] = (model, tfm)
    return model, tfm


# ===================== PATCH + BATCH =====================

def crop_patch(img: Image.Image, x: int, y: int, size: int = 32):
    half = size // 2
    left, top = x - half, y - half
    right, bottom = x + half, y + half

    patch = Image.new("RGB", (size, size), (0, 0, 0))
    crop = img.crop((left, top, right, bottom))

    paste_x = max(0, -left)
    paste_y = max(0, -top)
    patch.paste(crop, (paste_x, paste_y))
    return patch


@torch.no_grad()
def predict_points_batch(img, points, model, tfm):
    patches = [tfm(crop_patch(img, p["x"], p["y"])) for p in points]
    batch = torch.stack(patches)
    logits = model(batch)
    probs = torch.softmax(logits, dim=1)

    pred_idxs = probs.argmax(dim=1).cpu().numpy()
    confs = probs.max(dim=1).values.cpu().numpy()
    return pred_idxs, confs

def safe_sheet_name(name: str) -> str:
    """
    Excel: max 31 chars, no []:*?/\
    """
    bad = ['[', ']', ':', '*', '?', '/', '\\']
    for b in bad:
        name = name.replace(b, "_")
    name = name.strip() or "Imagen"
    return name[:31]

def safe_filename(name: str) -> str:
    name = os.path.basename(name or "imagen")
    return "".join(c if c.isalnum() or c in (" ", "-", "_", ".", "(", ")") else "_" for c in name).strip() or "imagen"

def add_points_to_sheet(ws, image_name: str, model_id: str, points: list):
    ws["A1"] = "imagen"
    ws["B1"] = image_name
    ws["A2"] = "modelo"
    ws["B2"] = model_id

    headers = ["idx", "label", "x", "y", "x_norm", "y_norm", "pred_label", "confidence", "source"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)

    for r, p in enumerate(points, start=5):
        ws.cell(r, 1, p.get("idx"))
        ws.cell(r, 2, p.get("label"))
        ws.cell(r, 3, p.get("x"))
        ws.cell(r, 4, p.get("y"))
        ws.cell(r, 5, p.get("x_norm"))
        ws.cell(r, 6, p.get("y_norm"))
        ws.cell(r, 7, p.get("pred_label"))
        ws.cell(r, 8, p.get("confidence"))
        ws.cell(r, 9, p.get("source"))

# ===================== ROUTE =====================

@app.post("/process")
async def process(
    file: UploadFile = File(...),
    n: int = Form(100),
    model_id: str = Form(""),
    points_json: str = Form("")  # 👈 NUEVO
):
    if not model_id:
        return JSONResponse({"error": "Selecciona un modelo"}, status_code=400)

    content = await file.read()
    img = Image.open(io.BytesIO(content)).convert("RGB")

    model, tfm = load_model_by_id(model_id)

    w, h = img.size

    # ✅ Si vienen puntos manuales, usarlos. Si no, generar random.
    points = []
    if points_json:
        try:
            points = json.loads(points_json)
        except Exception:
            return JSONResponse({"error": "points_json inválido"}, status_code=400)

        # Asegurar campos necesarios
        for i, p in enumerate(points, start=1):
            x = int(p.get("x", 0))
            y = int(p.get("y", 0))
            p["idx"] = int(p.get("idx", i))
            p["label"] = p.get("label") or idx_to_label_excel(p["idx"])
            p["x"] = x
            p["y"] = y
            p["x_norm"] = float(x / w) if w else 0.0
            p["y_norm"] = float(y / h) if h else 0.0
            p["source"] = p.get("source", "manual")
    else:
        points = generate_random_points(w, h, n=n)

    pred_idxs, confs = predict_points_batch(img, points, model, tfm)

    for p, idx, conf in zip(points, pred_idxs, confs):
        p["pred_label"] = CLASSES[int(idx)]
        p["confidence"] = float(conf)
        p.setdefault("source", "modelo")

    annotated = draw_points(img, points)

    buffer = io.BytesIO()
    annotated.save(buffer, format="PNG")
    img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return {
        "annotated_image_base64": img_base64,
        "points": points
    }


@app.get("/")
def home():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/models")
def list_models():
    items = []

    # Recorre todas las carpetas dentro de models_store
    for folder in os.listdir(MODELS_DIR):
        folder_path = os.path.join(MODELS_DIR, folder)
        if not os.path.isdir(folder_path):
            continue

        meta_path = os.path.join(folder_path, "meta.json")
        pt_path = os.path.join(folder_path, "best_model.pt")

        if not os.path.isfile(meta_path):
            continue

        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)

        # Asegura campos mínimos
        meta.setdefault("id", folder)
        meta.setdefault("model_name", folder)

        meta["classes"] = CLASSES
        meta["num_classes"] = NUM_CLASSES
        meta["has_weights"] = os.path.isfile(pt_path)

        items.append(meta)

    items.sort(key=lambda x: x.get("id", ""))
    return {"models": items}

@app.post("/export/excel")
def export_excel(payload: dict = Body(...)):
    points = payload.get("points", [])
    model_id = payload.get("model_id", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Puntos"

    # Encabezados
    headers = ["ID", "x", "y", "x_norm", "y_norm", "clase", "conf", "model_id"]
    ws.append(headers)

    # Filas
    for p in points:
        ws.append([
            p.get("label", ""),
            p.get("x", ""),
            p.get("y", ""),
            p.get("x_norm", ""),
            p.get("y_norm", ""),
            p.get("pred_label", ""),
            p.get("confidence", ""),
            model_id
        ])

    # Ajuste simple de ancho de columnas
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="puntos_coral.xlsx"'}
    )

@app.post("/process_batch")
async def process_batch(
    files: list[UploadFile] = File(...),
    n: int = Form(100),
    model_id: str = Form("")
):
    if not model_id:
        return JSONResponse({"error": "Selecciona un modelo"}, status_code=400)

    if not files or len(files) == 0:
        return JSONResponse({"error": "Sube al menos 1 imagen"}, status_code=400)

    if len(files) > 25:
        return JSONResponse({"error": "Máximo 25 imágenes"}, status_code=400)

    # Cargar modelo una vez
    model, tfm = load_model_by_id(model_id)

    # Excel workbook
    wb = Workbook()
    # Quitar la hoja por defecto
    default_ws = wb.active
    wb.remove(default_ws)

    # ZIP en memoria
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # procesar cada imagen
        used_sheet_names = set()

        for i, uf in enumerate(files, start=1):
            raw = await uf.read()
            try:
                img = Image.open(io.BytesIO(raw)).convert("RGB")
            except Exception:
                # si una imagen falla, la saltamos con una nota (opcional)
                continue

            w, h = img.size
            points = generate_random_points(w, h, n=n)

            pred_idxs, confs = predict_points_batch(img, points, model, tfm)
            for p, idx, conf in zip(points, pred_idxs, confs):
                p["pred_label"] = CLASSES[int(idx)]
                p["confidence"] = float(conf)
                p["source"] = "modelo"

            annotated = draw_points(img, points)

            # Nombre base
            original_name = safe_filename(uf.filename or f"imagen_{i}.png")
            base = os.path.splitext(original_name)[0]
            out_png_name = f"{base} (anotada).png"

            # Guardar PNG anotada en ZIP
            img_bytes = io.BytesIO()
            annotated.save(img_bytes, format="PNG")
            zf.writestr(f"imagenes_anotadas/{out_png_name}", img_bytes.getvalue())

            # Crear hoja Excel
            sheet_name = safe_sheet_name(base)
            # Evitar duplicados
            if sheet_name in used_sheet_names:
                k = 2
                while safe_sheet_name(f"{sheet_name}_{k}") in used_sheet_names:
                    k += 1
                sheet_name = safe_sheet_name(f"{sheet_name}_{k}")
            used_sheet_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            add_points_to_sheet(ws, original_name, model_id, points)

        # Guardar Excel dentro del ZIP
        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        zf.writestr("tabla_puntos.xlsx", xlsx_buf.getvalue())

    zip_buf.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"resultados_coral_{ts}.zip"

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
