import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def infer_method(p: dict) -> str:
    method = (p.get("method") or "").strip().lower()
    if method == "manual":
        return "manual"
    if method == "automatico":
        return "automatico"

    source = (p.get("source") or "").strip().lower()
    if source.startswith("manual"):
        return "manual"

    return "automatico"


def _write_points_sheet(ws, image_name: str, model_id: str, points: list):
    # Encabezado como tu screenshot
    ws["A1"] = "imagen"
    ws["B1"] = image_name
    ws["A2"] = "modelo"
    ws["B2"] = model_id

    headers = [
        "idx", "Etiqueta", "x", "y", "x_norm", "y_norm",
        "Predicción", "Confianza", "Método"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)

    for r, p in enumerate(points, start=5):
        ws.cell(r, 1, p.get("idx"))
        ws.cell(r, 2, p.get("label"))
        ws.cell(r, 3, p.get("x"))
        ws.cell(r, 4, p.get("y"))
        ws.cell(r, 5, p.get("x_norm"))
        ws.cell(r, 6, p.get("y_norm"))
        ws.cell(r, 7, p.get("pred_label"))
        ws.cell(r, 8, p.get("confidence"))
        ws.cell(r, 9, infer_method(p))

    widths = [8, 10, 10, 10, 14, 14, 20, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_single_excel(points: list, model_id: str, image_name: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Puntos"

    _write_points_sheet(ws, image_name=image_name, model_id=model_id, points=points)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def add_points_to_sheet(ws, image_name: str, model_id: str, points: list):
    # Batch usa esta función por hoja
    _write_points_sheet(ws, image_name=image_name, model_id=model_id, points=points)