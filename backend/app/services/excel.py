import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_single_excel(points: list, model_id: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Puntos"

    headers = ["ID", "x", "y", "x_norm", "y_norm", "clase", "conf", "model_id"]
    ws.append(headers)

    for p in points:
        ws.append([
            p.get("label", ""),
            p.get("x", ""),
            p.get("y", ""),
            p.get("x_norm", ""),
            p.get("y_norm", ""),
            p.get("pred_label", ""),
            p.get("confidence", ""),
            model_id
        ])

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def add_points_to_sheet(ws, image_name: str, model_id: str, points: list):
    ws["A1"] = "imagen"
    ws["B1"] = image_name
    ws["A2"] = "modelo"
    ws["B2"] = model_id

    headers = ["idx", "label", "x", "y", "x_norm", "y_norm", "pred_label", "confidence", "source"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)

    for r, p in enumerate(points, start=5):
        ws.cell(r, 1, p.get("idx"))
        ws.cell(r, 2, p.get("label"))
        ws.cell(r, 3, p.get("x"))
        ws.cell(r, 4, p.get("y"))
        ws.cell(r, 5, p.get("x_norm"))
        ws.cell(r, 6, p.get("y_norm"))
        ws.cell(r, 7, p.get("pred_label"))
        ws.cell(r, 8, p.get("confidence"))
        ws.cell(r, 9, p.get("source"))